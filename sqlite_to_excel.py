#!/usr/bin/env python3
"""
Convert SQLite database to Excel workbook.
Each table becomes a separate sheet.
"""

import sqlite3
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


def sqlite_to_excel(db_path, output_path):
    """Convert SQLite database to Excel file."""
    
    # Connect to SQLite database
    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()
    
    # Get all table names
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for table_name in tables:
        table_name = table_name[0]
        print(f"Processing table: {table_name}")
        
        # Get column names
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [row[1] for row in cursor.fetchall()]
        
        # Get all rows
        cursor.execute(f"SELECT * FROM {table_name};")
        rows = cursor.fetchall()
        
        # Create sheet
        ws = wb.create_sheet(title=table_name[:31])  # Excel sheet name limit is 31 chars
        
        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Save workbook
    wb.save(str(output_path))
    print(f"Excel file saved to: {output_path}")


if __name__ == "__main__":
    db_path = Path(__file__).parent / "db.sqlite"
    output_path = Path(__file__).parent / "database_export.xlsx"
    
    sqlite_to_excel(db_path, output_path)
